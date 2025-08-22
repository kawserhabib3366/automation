import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
import time
import logging
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import os
from urllib.parse import urlencode

cookies = {
    'ASP.NET_SessionId': 'slruu4552lgouv554jlya155',
    '__utma': '184618557.1420870435.1751483124.1751483124.1751483124.1',
    '__utmb': '184618557.5.10.1751483124',
    '__utmc': '184618557',
    '__utmz': '184618557.1751483124.1.1.utmcsr=(direct)^|utmccn=(direct)^|utmcmd=(none)',
    'cf_clearance': '5CAJYfbBBk193yM4egkwcYtUf6lvVoQcnnSlhor5maQ-1751483292-1.2.1.1-Ir4UYdi7wYCfmsmNoj7WNP2bZPogtpcXFKIkFWUP6OlcmHWOmNQT4HP6xmG2k1GwiZsD24z.7wVlGg7OJxe82e2_g4Oyc9fPcgpd9GGk1pglfaFvORl6NiWG3eUAGZN_moqxjbAsKvNQ6TYW7A0AWwTlfDdrW7LiwwwhrTI4FK6PIKl3gDhkdbtZpKRKhruM7NydX101f6cQ7pdkJMLBEzYbt63r93zhw7NZF8bF2.Q',
    '_ga_ZRGXL7XFHK': 'GS2.1.s1751483295$o1$g0$t1751483295$j60$l0$h0',
    '_ga': 'GA1.4.1327291438.1751483295',
    '_gid': 'GA1.4.889351127.1751483296',
    '_fbp': 'fb.3.1751483296844.763051666818148596',
}

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:136.0) Gecko/20100101 Firefox/136.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    # 'Accept-Encoding': 'gzip, deflate, br, zstd',
    'Content-Type': 'application/x-www-form-urlencoded',
    'Origin': 'https://eservices.boroondara.vic.gov.au',
    'Connection': 'keep-alive',
    'Referer': 'https://eservices.boroondara.vic.gov.au/EPlanning/Pages/XC.Track/SearchApplication.aspx?d=lastmonth&k=DeterminationDate&t=PlnPermit,PlnAppeals,PlnPostPer,PlanPermGr,PlanAmend,PlanAppeal',
    # 'Cookie': 'ASP.NET_SessionId=slruu4552lgouv554jlya155; __utma=184618557.1420870435.1751483124.1751483124.1751483124.1; __utmb=184618557.5.10.1751483124; __utmc=184618557; __utmz=184618557.1751483124.1.1.utmcsr=(direct)^|utmccn=(direct)^|utmcmd=(none); cf_clearance=5CAJYfbBBk193yM4egkwcYtUf6lvVoQcnnSlhor5maQ-1751483292-1.2.1.1-Ir4UYdi7wYCfmsmNoj7WNP2bZPogtpcXFKIkFWUP6OlcmHWOmNQT4HP6xmG2k1GwiZsD24z.7wVlGg7OJxe82e2_g4Oyc9fPcgpd9GGk1pglfaFvORl6NiWG3eUAGZN_moqxjbAsKvNQ6TYW7A0AWwTlfDdrW7LiwwwhrTI4FK6PIKl3gDhkdbtZpKRKhruM7NydX101f6cQ7pdkJMLBEzYbt63r93zhw7NZF8bF2.Q; _ga_ZRGXL7XFHK=GS2.1.s1751483295$o1$g0$t1751483295$j60$l0$h0; _ga=GA1.4.1327291438.1751483295; _gid=GA1.4.889351127.1751483296; _fbp=fb.3.1751483296844.763051666818148596',
    'Upgrade-Insecure-Requests': '1',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-User': '?1',
    'Priority': 'u=0, i',
}


BASE_URL = "https://eservices.boroondara.vic.gov.au/EPlanning/Pages/XC.Track/SearchApplication.aspx?d=lastmonth&k=DeterminationDate&t=PlnPermit%2cPlnAppeals%2cPlnPostPer%2cPlanPermGr%2cPlanAmend%2cPlanAppeal"
session = requests.Session()

EXCEL_FILE = "churches.xlsx"

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

def get_hidden_fields(soup):
    return {
        key: (soup.find("input", {"name": key})["value"] if soup.find("input", {"name": key}) else "")
        for key in ["__VIEWSTATE", "__VIEWSTATEGENERATOR", "__EVENTVALIDATION"]
    }


def scrape_church_blocks(soup) -> pd.DataFrame:
    churches = []

    for name_span in soup.find_all("span", id=re.compile(r"_churches_name_")):
        table = name_span.find_parent("table")

        def grab(pattern):
            tag = table.find("span", id=re.compile(pattern))
            return tag.get_text(strip=True) if tag else ""

        full_name = grab(r"_churches_pastor_")
        name_parts = full_name.strip().split()
        address = grab(r"_addressLine2_")
        #print(address)

        # Clean the address
        address = address.strip().replace('\xa0', ' ')  # remove non-breaking spaces
        address = re.sub(r'\s+', ' ', address)  # normalize multiple spaces to one

        # Define regex pattern
        pattern = r"^(.*?),\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)$"

        # Match
        match = re.match(pattern, address)

        if match:
            #print("found match:", repr(address))  # helpful for debugging
            city = match.group(1)
            state = match.group(2)
            zip_code = match.group(3)
        else:
            print("No match found:", repr(address))  # helpful for debugging
            break


        church = {
            "Church Name": name_span.get_text(strip=True),
            "First Name": name_parts[0] if len(name_parts) >= 1 else "",
            "Last Name": " ".join(name_parts[1:]) if len(name_parts) >= 2 else "",
            "Address Line 1": grab(r"_addressLine1_"),
            "city": city,
            "state": state,
            "zip_code": zip_code,
            "Phone": grab(r"_phoneNo_"),
            "Email": grab(r"_email_"),
            "Homepage": (table.find("a", id=re.compile(r"_url_"))["href"]
                         if table.find("a", id=re.compile(r"_url_")) else ""),
            "Map Link": (table.find("a", id=re.compile(r"_mapLink_"))["href"]
                         if table.find("a", id=re.compile(r"_mapLink_")) else ""),
            "Schedule": " | ".join(" ".join(li.stripped_strings) for li in
                                   (table.find("div", id=re.compile(r"_schedules_list_")) or []).find_all("li"))
            if table.find("div", id=re.compile(r"_schedules_list_")) else ""
        }

        churches.append(church)

    return pd.DataFrame(churches)

def write_df_to_excel(df: pd.DataFrame, filename: str, sheet_name: str = "Sheet1"):
    """Append a DataFrame to an Excel file (create if not exists)."""
    if not os.path.exists(filename):
        # Create file and write headers
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        wb.save(filename)
    else:
        # Append without headers
        wb = load_workbook(filename)
        ws = wb[sheet_name]
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
        wb.save(filename)

def paginate_and_scrape():
    resp = session.get(BASE_URL,
    cookies=cookies,
    headers=headers)
    soup = BeautifulSoup(resp.text, "html.parser")





    page = 1
    for each in range(20):
        logging.info(f"Scraping Page {page}")
        # df = scrape_church_blocks(soup)
        # if df.empty:
        #     logging.info("No more data found.")
        #     break

        # write_df_to_excel(df, EXCEL_FILE)

        # Find Next Page button
        # next_btn = soup.find("a", id=lambda x: x and "c_pageNext" in x)

        # if not next_btn:
        #     logging.info("No more pages to scrape (no next button found).")
        #     break

        # # Check if the button is disabled by looking for the class 'aspNetDisabled'
        # if "aspNetDisabled" in next_btn.get("class", ''):
        #     logging.info("Next button is disabled, no more pages.")
        #     break

        # href = next_btn.get("href")
        # if not href:
        #     logging.info("Next button has no href attribute.")
        #     break

        payload_dict = {
            "__EVENTTARGET": "pager",
            "__EVENTARGUMENT": "10|FgxttsZfhiDdjCoKuqiT",  # This second part changes per session – scrape it if dynamic
            "__LASTFOCUS": "",
            **get_hidden_fields(soup),
            "Master$txtSearch": "",
            "Master$ctMain1$apsearch$txtSearch$txtSearch_txt1": "",
        }
         
        #payload = urlencode(payload_dict)

        
        

        try:
            #print(payload_dict)

            resp = session.post(BASE_URL,cookies=cookies,
    headers=headers,data=payload_dict)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.text, "html.parser")
            page += 1
            if page ==2:
                print(soup)
                break
            time.sleep(1.5)
        except Exception as e:
            logging.error(f"Failed to load page {page + 1}: {e}")
            break

def main():

    paginate_and_scrape()
    logging.info(f"Data saved to {EXCEL_FILE}")

if __name__ == "__main__":
    main()
